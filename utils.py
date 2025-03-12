import pandas as pd
import re
import matplotlib as plt
import pandas as pd
import numpy as np
import datetime
import calendar
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import numbers
from datetime import datetime
import pandas as pd
import sys
from sys import argv
import os
import numpy as np
from itertools import product
from datetime import datetime
from datetime import timedelta
from pandas.tseries.offsets import MonthEnd
from dateutil.relativedelta import relativedelta
import shutil
import io
import json
import gspread
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
from gspread_dataframe import set_with_dataframe
from pathlib import Path
import re
import trino
import trino.auth
import streamlit as st
import base64
from email.mime.text import MIMEText
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import os
import pickle
import base64
import mimetypes
from googleapiclient.discovery import build
from email.message import EmailMessage
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.errors import HttpError
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def connection_starburst():
    """Starbust function connector."""
    import trino
    import trino.auth
    try:
        conn = trino.dbapi.connect(
            host='starburst.g8s-data-platform-prod.glovoint.com',
            port=443,
            http_scheme='https',
            auth=trino.auth.OAuth2Authentication()
        )
        cursor = conn.cursor()
    except Exception as e:
        print(f'Error connecting to Starburst: {e}')
        raise
    return conn

def read_sheet(creds, spreadsheet_id_fx: str, sheet_range: str):
    """Reads a specific spreadsheet and returns the data in a pd.DataFrame.
    
    Args:
        creds: Google API credentials.
        spreadsheet_id_fx: Google Workbook ID.
        sheet_range: Range of the sheet to extract (e.g., Sheet1!A:B).
        
    Returns:
        df_to_return: pd.DataFrame with the spreadsheet data.
    """
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=spreadsheet_id_fx, range=sheet_range).execute()
    values = result.get('values', [])
    df_to_return = pd.DataFrame(values)
    df_to_return.columns = df_to_return.iloc[0, :]
    df_to_return = df_to_return.drop(labels=0, axis=0)
    return df_to_return

def write_sheet(creds, spreadsheet_id: str, sheet_name: str, df_to_write, start_row=3):
    """Writes to a specific spreadsheet based on the provided pd.DataFrame.
    
    Args:
        creds: Google API credentials.
        spreadsheet_id: Google Workbook ID.
        sheet_name: Name of the sheet to write to.
        df_to_write: pd.DataFrame with the data to write.
        start_row: Starting row to write the data (default is 3).
    """
    gs = gspread.authorize(creds)
    worksheet = gs.open_by_key(spreadsheet_id).worksheet(sheet_name)

    # Convertir el DataFrame en una lista de listas
    data = df_to_write.values.tolist()

    # Escribir los datos comenzando en la celda especificada
    cell_list = worksheet.range(f"A{start_row}:{get_column_letter(len(df_to_write.columns))}{start_row + len(data) - 1}")
    for cell, value in zip(cell_list, data):
        cell.value = value
    worksheet.update_cells(cell_list)

def extract_invoice_numbers(df):
    """Extracts invoice numbers from the Description column of the DataFrame.
    
    Args:
        df: pd.DataFrame with the descriptions.
        
    Returns:
        result_df: pd.DataFrame with invoice numbers and amounts.
    """
    # Verificar si las columnas esperadas existen en el DataFrame
    expected_columns = ['Description', 'Flow amount']
    missing_columns = [col for col in expected_columns if col not in df.columns]
    if missing_columns:
        st.error(f"The following columns are missing in the Excel file: {missing_columns}")
        return pd.DataFrame()  # Retornar un DataFrame vacío si faltan columnas
    
    # Normalizar los nombres de las columnas para evitar problemas de mayúsculas/minúsculas o espacios adicionales
    df.columns = df.columns.str.strip().str.lower()

    # Inicializar la lista para almacenar los resultados
    results = []

    # Iterar sobre las filas del DataFrame
    for index, row in df.iterrows():
        description = row['description']
        flow_amount = row['flow amount']

        # Buscar todos los números que comienzan con '22' y tienen 10 caracteres, eliminando duplicados
        matches = list(set(re.findall(r'\b22\d{8}\b', str(description))))

        # Añadir cada número encontrado al resultado con su correspondiente 'Flow Amount'
        for match in matches:
            results.append({'Invoice Number': match, 'Flow Amount': flow_amount})

    # Convertir los resultados a un DataFrame
    if results:
        result_df = pd.DataFrame(results)
        st.write("Extracted invoice numbers:")
        st.write(result_df.head())
    else:
        st.warning("No invoice numbers were found in the 'Description' column.")
        result_df = pd.DataFrame()

    return result_df

# SCOPES para Gmail
GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']

# Autenticación de Gmail usando OAuth2.0
def gmail_authenticate(creds_file):
    """Authenticate with Gmail using OAuth 2.0."""
    creds = None
    if os.path.exists('gmail_token.pickle'):
        with open('gmail_token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(creds_file, GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open('gmail_token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    return creds

# Enviar correo usando Gmail API

# Enviar correo usando Gmail API
def send_email(creds, to, subject, body, attachment_path=None, is_html=False):
    """Crear y enviar un mensaje de correo electrónico con adjunto, soportando texto plano y HTML."""
    try:
        # Crear cliente de la API de Gmail
        service = build('gmail', 'v1', credentials=creds)
        message = EmailMessage()

        # Verificar si el cuerpo del mensaje es HTML o texto plano
        if is_html:
            message.set_content(body, subtype='html')
        else:
            message.set_content(body, subtype='plain')

        message['To'] = to
        message['From'] = 'bernat.morros@glovoapp.com'
        message['Subject'] = subject

        # Adjuntar el archivo instructivo por defecto
        # default_attachment_path = os.path.join("processes", "attachments", "Instructivo_validacion_de_factura.pdf")
        # with open(default_attachment_path, 'rb') as f:
        #     content = f.read()
        # maintype, subtype = mimetypes.guess_type(default_attachment_path)[0].split('/')
        # message.add_attachment(content, maintype, subtype, filename=os.path.basename(default_attachment_path))

        # Si hay un archivo adicional adjunto, se adjunta también
        if attachment_path:
            attachment_filename = os.path.basename(attachment_path)
            with open(attachment_path, 'rb') as f:
                content = f.read()
            maintype, subtype = mimetypes.guess_type(attachment_filename)[0].split('/')
            message.add_attachment(content, maintype, subtype, filename=attachment_filename)

        # Codificar el mensaje
        encoded_message = base64.urlsafe_b64encode(message.as_bytes()).decode()

        create_message = {
            'raw': encoded_message
        }

        # Enviar el mensaje
        send_message = service.users().messages().send(userId="me", body=create_message).execute()
        print(f'Message Id: {send_message["id"]}')
        return send_message
    except HttpError as error:
        print(f'An error occurred: {error}')
        return None
    
# Función para validar el formato de un correo electrónico
def es_email_valido(email):
    # Regex simple para validar emails
    patron = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
    return re.match(patron, email) is not None
